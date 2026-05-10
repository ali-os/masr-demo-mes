#!/usr/bin/env python3
"""
MASR MES — Full Excel Ingestion Engine
Supports both Daily Prod Report (.xlsb) and OEE (.xlsx) files.
Reads MPS, W-1/W-2/W-3 weekly schedules, and daily actuals.
Outputs JSON results and upserts into PostgreSQL.
"""
import sys
import json
import argparse
import re
from datetime import datetime, date

# Try importing Excel libs
try:
    from pyxlsb import open_workbook as open_xlsb
    HAS_XLSB = True
except ImportError:
    HAS_XLSB = False

try:
    import openpyxl
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

try:
    import psycopg2
    HAS_PG = True
except ImportError:
    HAS_PG = False


def clean_val(v):
    if v is None:
        return None
    if isinstance(v, float) and v != v:  # NaN check
        return None
    return v

def to_int(v):
    try:
        return int(float(v)) if v else 0
    except:
        return 0

def to_float(v):
    try:
        return float(v) if v else 0.0
    except:
        return 0.0

def normalize_sku(v):
    """Convert numeric SKU to 8-digit string"""
    if not v:
        return None
    try:
        return str(int(float(v)))
    except:
        return str(v).strip()


def read_xlsb_sheet(wb, sheet_name):
    rows = []
    try:
        with wb.get_sheet(sheet_name) as sheet:
            for row in sheet.rows():
                rows.append([clean_val(c.v) for c in row])
    except Exception as e:
        pass
    return rows


def read_xlsx_sheet(wb, sheet_name):
    rows = []
    try:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            rows.append([clean_val(c) for c in row])
    except Exception as e:
        pass
    return rows


def parse_mps_sheet(rows, year, month):
    """Parse the MPS sheet: Code|Brand|Family|Description|Size|Category|Line|MPS|Bulk"""
    plans = []
    header_found = False
    for row in rows:
        # Find header row
        row_str = ' '.join(str(c) for c in row if c)
        if 'Code' in row_str and 'Brand' in row_str and 'Family' in row_str:
            header_found = True
            continue
        if not header_found:
            continue
        if len(row) < 8:
            continue
        sku = normalize_sku(row[0])
        if not sku or len(sku) < 6:
            continue
        plans.append({
            'skuCode': sku,
            'brand': str(row[1]).strip() if row[1] else '',
            'family': str(row[2]).strip() if row[2] else '',
            'category': str(row[4]).strip() if len(row) > 4 and row[4] else '',
            'machineName': str(row[6]).strip() if len(row) > 6 and row[6] else '',
            'targetQty': to_int(row[7]) if len(row) > 7 else 0,
            'bulkKg': to_float(row[8]) if len(row) > 8 else 0.0,
            'year': year,
            'month': month
        })
    return plans


def parse_weekly_sheet(rows, week_no, year, month):
    """Parse W-1/W-2/W-3: SKU|Brand|Family|Name|Volume|Machine|WeekTarget|Bulk|Day1Sh1|Day1Sh2..."""
    entries = []
    day_dates = []
    header_found = False

    for i, row in enumerate(rows):
        if not row:
            continue
        row_str = ' '.join(str(c) for c in row if c)

        # Find date header row (numeric date codes like 46082.0)
        if i > 0 and not header_found:
            date_vals = [c for c in row if isinstance(c, (int, float)) and c and 46000 < float(c) < 47000]
            if len(date_vals) >= 3:
                # Convert Excel date serial to real date
                from datetime import datetime, timedelta
                excel_epoch = datetime(1899, 12, 30)
                day_dates = [(excel_epoch + timedelta(days=int(float(d)))).date() for d in date_vals]
                header_found = True
                continue

        if not header_found:
            continue

        sku = normalize_sku(row[0]) if row[0] else None
        if not sku or len(sku) < 6:
            continue

        week_target = to_int(row[6]) if len(row) > 6 else 0
        if week_target == 0:
            continue

        # Extract per-day per-shift actuals
        # Columns 8+ are day1_sh1, day1_sh2, day2_sh1, day2_sh2, ...
        for d_idx, entry_date in enumerate(day_dates):
            col_sh1 = 8 + d_idx * 2
            col_sh2 = 9 + d_idx * 2
            sh1_qty = to_int(row[col_sh1]) if len(row) > col_sh1 else 0
            sh2_qty = to_int(row[col_sh2]) if len(row) > col_sh2 else 0

            if sh1_qty > 0:
                entries.append({
                    'skuCode': sku,
                    'machineName': str(row[5]).strip() if len(row) > 5 and row[5] else '',
                    'date': entry_date.isoformat(),
                    'shift': 1,
                    'actualQty': sh1_qty,
                    'weekTarget': week_target,
                    'year': year,
                    'month': month
                })
            if sh2_qty > 0:
                entries.append({
                    'skuCode': sku,
                    'machineName': str(row[5]).strip() if len(row) > 5 and row[5] else '',
                    'date': entry_date.isoformat(),
                    'shift': 2,
                    'actualQty': sh2_qty,
                    'weekTarget': week_target,
                    'year': year,
                    'month': month
                })

    return entries


def ingest_to_db(db_url, plans, entries, year, month):
    """Upsert plans and daily entries into PostgreSQL with Transaction support"""
    if not HAS_PG:
        return {'inserted': 0, 'skipped': 0, 'warnings': ['psycopg2 not installed, output JSON only']}

    conn = psycopg2.connect(db_url)
    inserted = 0
    skipped = 0
    warnings = []

    try:
        cur = conn.cursor()
        
        # 1. Start Transaction (implicit in psycopg2, but we'll be explicit about rollback)
        
        # Upsert MPS plans
        for p in plans:
            cur.execute('SELECT id FROM "Product" WHERE "skuCode" = %s', (p['skuCode'],))
            row = cur.fetchone()
            if not row:
                warnings.append(f"SKU not found: {p['skuCode']}")
                skipped += 1
                continue
            product_id = row[0]

            cur.execute('''
                INSERT INTO "MonthlyPlan" (id, year, month, "productId", "machineName", "targetQty", "bulkKg", category, status, "createdAt", "updatedAt")
                VALUES (gen_random_uuid(), %s, %s, %s, %s, %s, %s, %s, 'DRAFT'::"PlanStatus", NOW(), NOW())
                ON CONFLICT (year, month, "productId") DO UPDATE
                SET "machineName" = EXCLUDED."machineName",
                    "targetQty" = EXCLUDED."targetQty",
                    "bulkKg" = EXCLUDED."bulkKg",
                    category = EXCLUDED.category,
                    "updatedAt" = NOW()
            ''', (p['year'], p['month'], product_id, p['machineName'], p['targetQty'], p['bulkKg'], p['category']))
            inserted += 1

        # Upsert daily entries
        for e in entries:
            cur.execute('SELECT id FROM "Product" WHERE "skuCode" = %s', (e['skuCode'],))
            prod_row = cur.fetchone()
            if not prod_row:
                skipped += 1
                continue
            product_id = prod_row[0]

            cur.execute('SELECT id FROM "MonthlyPlan" WHERE year = %s AND month = %s AND "productId" = %s',
                        (e['year'], e['month'], product_id))
            plan_row = cur.fetchone()
            if not plan_row:
                skipped += 1
                continue
            plan_id = plan_row[0]

            cur.execute('''
                INSERT INTO "DailyEntry" (id, "monthlyPlanId", "productId", date, shift, "actualQty", "createdAt", "updatedAt")
                VALUES (gen_random_uuid(), %s, %s, %s::date, %s, %s, NOW(), NOW())
                ON CONFLICT ("monthlyPlanId", date, shift) DO UPDATE
                SET "actualQty" = EXCLUDED."actualQty", "updatedAt" = NOW()
            ''', (plan_id, product_id, e['date'], e['shift'], e['actualQty']))
            inserted += 1

        conn.commit()
        cur.close()
    except Exception as ex:
        conn.rollback()
        return {'inserted': 0, 'skipped': 0, 'error': str(ex), 'warnings': warnings}
    finally:
        conn.close()

    return {'inserted': inserted, 'skipped': skipped, 'warnings': warnings}


def main():
    parser = argparse.ArgumentParser(description='MASR Excel Importer')
    parser.add_argument('--file', required=True, help='Path to .xlsb or .xlsx file')
    parser.add_argument('--type', default='DAILY_PROD', choices=['DAILY_PROD', 'OEE'])
    parser.add_argument('--year', type=int, default=2026)
    parser.add_argument('--month', type=int, default=3)
    parser.add_argument('--db-url', default=None, help='PostgreSQL connection URL')
    args = parser.parse_args()

    filepath = args.file
    is_xlsb = filepath.lower().endswith('.xlsb')

    plans = []
    entries = []

    if is_xlsb:
        if not HAS_XLSB:
            print(json.dumps({'error': 'pyxlsb not installed', 'inserted': 0, 'skipped': 0}))
            sys.exit(1)
        with open_xlsb(filepath) as wb:
            sheet_names = wb.sheets
            # MPS
            for sn in sheet_names:
                if 'MPS' in sn.upper():
                    rows = read_xlsb_sheet(wb, sn)
                    plans.extend(parse_mps_sheet(rows, args.year, args.month))
            # Weekly actuals
            for week_no, sn in enumerate(['W-1', 'W-2', 'W-3'], 1):
                if sn in sheet_names:
                    rows = read_xlsb_sheet(wb, sn)
                    entries.extend(parse_weekly_sheet(rows, week_no, args.year, args.month))
    else:
        if not HAS_XLSX:
            print(json.dumps({'error': 'openpyxl not installed', 'inserted': 0, 'skipped': 0}))
            sys.exit(1)
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        for sn in wb.sheetnames:
            if 'MPS' in sn.upper():
                rows = read_xlsx_sheet(wb, sn)
                plans.extend(parse_mps_sheet(rows, args.year, args.month))
            for week_no, pat in enumerate(['W-1', 'W-2', 'W-3'], 1):
                if pat.upper() in sn.upper():
                    rows = read_xlsx_sheet(wb, sn)
                    entries.extend(parse_weekly_sheet(rows, week_no, args.year, args.month))

    # Deduplicate plans by skuCode (keep last)
    plan_map = {}
    for p in plans:
        plan_map[p['skuCode']] = p
    plans = list(plan_map.values())

    result = {'plans_found': len(plans), 'entries_found': len(entries), 'inserted': 0, 'skipped': 0, 'warnings': []}

    if args.db_url:
        db_result = ingest_to_db(args.db_url, plans, entries, args.year, args.month)
        result.update(db_result)
    else:
        result['data'] = {'plans': plans[:5], 'entries': entries[:5]}
        result['warnings'].append('No DB URL provided, dry run only')

    print(json.dumps(result))


if __name__ == '__main__':
    main()
